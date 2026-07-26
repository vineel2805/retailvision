"""Report Exporter for CSV, Excel (.xlsx), and PDF formats (FR-010)."""

from __future__ import annotations

import csv
import io
from datetime import date
from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


class ReportExporter:
    """Generates downloadable CSV, XLSX, and PDF files for visitor report summaries."""

    @staticmethod
    def to_csv(summary_data: Dict[str, Any]) -> str:
        """Generate CSV string from report summary data."""
        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow(["RetailVision Footfall Report"])
        writer.writerow(["Period", f"{summary_data.get('start_date')} to {summary_data.get('end_date')}"])
        writer.writerow(["Total Entries", summary_data.get("total_entries", 0)])
        writer.writerow(["Total Exits", summary_data.get("total_exits", 0)])
        writer.writerow(["Peak Hour", summary_data.get("peak_hour", "N/A")])
        writer.writerow([])

        writer.writerow(["Date", "Hour", "Entries", "Exits", "Peak Hour"])
        for row in summary_data.get("hourly_breakdown", []):
            writer.writerow([
                row.get("event_date"),
                f"{row.get('hour')}:00",
                row.get("entries"),
                row.get("exits"),
                "YES" if row.get("is_peak_hour") else "NO",
            ])

        return output.getvalue()

    @staticmethod
    def to_xlsx(summary_data: Dict[str, Any]) -> bytes:
        """Generate Excel workbook bytes using openpyxl."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Footfall Report"

        header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        title_font = Font(name="Arial", size=14, bold=True)

        ws.append(["RetailVision Footfall Report"])
        ws.cell(row=1, column=1).font = title_font
        ws.append(["Period", f"{summary_data.get('start_date')} to {summary_data.get('end_date')}"])
        ws.append(["Total Entries", summary_data.get("total_entries", 0)])
        ws.append(["Total Exits", summary_data.get("total_exits", 0)])
        ws.append(["Peak Hour", summary_data.get("peak_hour", "N/A")])
        ws.append([])

        headers = ["Date", "Hour", "Entries", "Exits", "Peak Hour"]
        ws.append(headers)
        header_row_idx = ws.max_row

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row in summary_data.get("hourly_breakdown", []):
            ws.append([
                row.get("event_date"),
                f"{row.get('hour')}:00",
                row.get("entries"),
                row.get("exits"),
                "YES" if row.get("is_peak_hour") else "NO",
            ])

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def to_pdf(summary_data: Dict[str, Any]) -> bytes:
        """Generate PDF document bytes using reportlab."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "TitleStyle",
            parent=styles["Heading1"],
            fontSize=18,
            textColor=colors.HexColor("#0F172A"),
            spaceAfter=12,
        )
        meta_style = ParagraphStyle(
            "MetaStyle",
            parent=styles["Normal"],
            fontSize=10,
            textColor=colors.HexColor("#475569"),
            spaceAfter=6,
        )

        elements = []
        elements.append(Paragraph("RetailVision Footfall Report", title_style))
        elements.append(
            Paragraph(
                f"<b>Date Range:</b> {summary_data.get('start_date')} to {summary_data.get('end_date')} | "
                f"<b>Total Entries:</b> {summary_data.get('total_entries', 0)} | "
                f"<b>Total Exits:</b> {summary_data.get('total_exits', 0)} | "
                f"<b>Peak Hour:</b> {summary_data.get('peak_hour', 'N/A')}",
                meta_style,
            )
        )
        elements.append(Spacer(1, 12))

        table_data = [["Date", "Hour", "Entries", "Exits", "Peak Hour"]]
        for row in summary_data.get("hourly_breakdown", []):
            table_data.append([
                str(row.get("event_date")),
                f"{row.get('hour')}:00",
                str(row.get("entries")),
                str(row.get("exits")),
                "YES" if row.get("is_peak_hour") else "NO",
            ])

        t = Table(table_data, colWidths=[110, 80, 80, 80, 90])
        t.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ])
        )

        elements.append(t)
        doc.build(elements)
        return buffer.getvalue()
